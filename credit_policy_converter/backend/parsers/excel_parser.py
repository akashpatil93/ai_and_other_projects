"""Excel document parser for credit policy documents."""
import openpyxl
from typing import List, Dict, Any


def parse_excel(file_path: str) -> List[Dict[str, Any]]:
    """
    Parse an Excel file and return all sheets as structured sections.
    Each section has: name, headers, rows (list of dicts), text (for LLM), row_count.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sections = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Collect all non-empty rows
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                all_rows.append(row)

        if not all_rows:
            continue

        # Find the header row (first row with 2+ non-empty cells)
        header_idx = 0
        for i, row in enumerate(all_rows):
            non_empty = sum(1 for c in row if c is not None and str(c).strip())
            if non_empty >= 2:
                header_idx = i
                break

        # Build headers list, replacing blank cells with placeholder names
        raw_headers = all_rows[header_idx]
        headers = []
        for j, cell in enumerate(raw_headers):
            if cell is not None and str(cell).strip():
                headers.append(str(cell).strip())
            else:
                headers.append(f"column_{j + 1}")

        # Extract data rows as dicts
        data_rows = []
        for row in all_rows[header_idx + 1:]:
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                row_dict = {}
                for j, cell in enumerate(row):
                    if j < len(headers):
                        row_dict[headers[j]] = cell
                data_rows.append(row_dict)

        # Build plain-text representation for LLM context
        text_parts = [f"=== Sheet: {sheet_name} ===", " | ".join(headers), "-" * 80]
        for row in data_rows[:150]:
            text_parts.append(" | ".join(
                str(row.get(h, "")) if row.get(h) is not None else ""
                for h in headers
            ))
        if len(data_rows) > 150:
            text_parts.append(f"... ({len(data_rows) - 150} more rows omitted)")

        sections.append({
            "name": sheet_name,
            "headers": headers,
            "rows": data_rows,
            "text": "\n".join(text_parts),
            "row_count": len(data_rows),
        })

    wb.close()
    return sections
